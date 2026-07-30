"""Export des résultats d'extraction / scoring (JSON téléchargeable + Excel)."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def build_json_bytes(payload: dict[str, Any]) -> tuple[bytes, str]:
    """Sérialise le résultat calculé en JSON téléchargeable."""
    body = json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode(
        "utf-8"
    )
    return body, f"analyse_financiere_{_ts()}.json"


def _autofit(ws, max_width: int = 48) -> None:
    for col_idx, column in enumerate(ws.columns, start=1):
        length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, length + 2)


def _header_row(ws, headers: list[str], fill: str = "F58220") -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=fill)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _cell(value: Any) -> Any:
    if value is None or isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def _write_financial_analysis_sheets(
    wb: Workbook,
    *,
    extraction: dict[str, Any],
    analysis: dict[str, Any],
    sheet_prefix: str = "",
) -> None:
    """Feuilles Excel pour un document analyse financière (qwen_only)."""
    prefix = sheet_prefix or ""

    def _title(name: str) -> str:
        return f"{prefix}{name}"[:31]

    ws = wb.create_sheet(_title("Synthese"))
    _header_row(ws, ["Champ", "Valeur"])
    decision = (analysis or {}).get("decision") or {}
    rows = [
        ("source_filename", extraction.get("source_filename")),
        ("pages_total", extraction.get("pages_total")),
        ("pages_ok", extraction.get("pages_ok")),
        ("pages_failed", extraction.get("pages_failed")),
        ("processing_time_ms", extraction.get("processing_time_ms")),
        ("model", extraction.get("model")),
        ("scoring_mode", (analysis or {}).get("scoring_mode")),
        ("final_score", (analysis or {}).get("final_score")),
        ("decision_score", decision.get("score")),
        ("risk_class", decision.get("risk_class")),
        ("decision", decision.get("decision")),
        ("recommendation", decision.get("recommendation")),
        ("blocking_status", decision.get("blocking_status")),
        ("profile", decision.get("profile")),
    ]
    mapping = (analysis or {}).get("mapping") or {}
    if mapping:
        rows.extend(
            [
                ("mapping_strategy", mapping.get("strategy")),
                ("mapping_model", mapping.get("model")),
                ("sections_processed", mapping.get("sections_processed")),
                ("candidates_total", mapping.get("candidates_total")),
            ]
        )
    for i, (key, value) in enumerate(rows, start=2):
        ws.cell(i, 1, key)
        ws.cell(i, 2, _cell(value))
    _autofit(ws)

    dataset = (analysis or {}).get("dataset") or {}
    ws_ds = wb.create_sheet(_title("Dataset"))
    _header_row(ws_ds, ["Code", "Label", "Valeur", "Unité", "Statut"])
    row = 2
    for key, field in dataset.items():
        if not isinstance(field, dict) or "status" not in field:
            continue
        ws_ds.cell(row, 1, field.get("code") or key)
        ws_ds.cell(row, 2, field.get("label") or key)
        ws_ds.cell(row, 3, _cell(field.get("value")))
        ws_ds.cell(row, 4, field.get("unit"))
        ws_ds.cell(row, 5, field.get("status"))
        row += 1
    _autofit(ws_ds)

    ratios = (analysis or {}).get("ratios") or []
    ws_r = wb.create_sheet(_title("Ratios"))
    _header_row(
        ws_r,
        ["Code", "Label", "Valeur", "Unité", "Statut", "Seuil", "Points", "Max", "Formule"],
    )
    if isinstance(ratios, list):
        for i, ratio in enumerate(ratios, start=2):
            ws_r.cell(i, 1, ratio.get("code"))
            ws_r.cell(i, 2, ratio.get("label"))
            ws_r.cell(i, 3, _cell(ratio.get("value")))
            ws_r.cell(i, 4, ratio.get("unit"))
            ws_r.cell(i, 5, ratio.get("status"))
            ws_r.cell(i, 6, ratio.get("threshold"))
            ws_r.cell(i, 7, _cell(ratio.get("points")))
            ws_r.cell(i, 8, _cell(ratio.get("max_points")))
            ws_r.cell(i, 9, ratio.get("formula"))
    _autofit(ws_r)

    axes = (analysis or {}).get("axes") or []
    ws_a = wb.create_sheet(_title("Axes"))
    _header_row(
        ws_a,
        ["Code", "Label", "Score", "Poids", "Contribution", "Calculable"],
    )
    if isinstance(axes, list):
        for i, axe in enumerate(axes, start=2):
            ws_a.cell(i, 1, axe.get("code"))
            ws_a.cell(i, 2, axe.get("label"))
            ws_a.cell(i, 3, _cell(axe.get("raw_score")))
            ws_a.cell(i, 4, _cell(axe.get("weight")))
            ws_a.cell(i, 5, _cell(axe.get("weighted_contribution")))
            ws_a.cell(i, 6, axe.get("calculable"))
    _autofit(ws_a)

    checks = (analysis or {}).get("accounting_checks") or []
    if checks:
        ws_c = wb.create_sheet(_title("Controles"))
        _header_row(ws_c, ["Code", "Statut", "Attendu", "Observé", "Écart", "Message"])
        for i, check in enumerate(checks, start=2):
            ws_c.cell(i, 1, check.get("code"))
            ws_c.cell(i, 2, check.get("status"))
            ws_c.cell(i, 3, _cell(check.get("expected")))
            ws_c.cell(i, 4, _cell(check.get("observed")))
            ws_c.cell(i, 5, _cell(check.get("difference")))
            ws_c.cell(i, 6, check.get("message"))
        _autofit(ws_c)


def _write_legacy_liasse_sheets(
    wb: Workbook,
    payload: dict[str, Any],
) -> None:
    extraction = payload.get("extraction") or payload
    scoring = payload.get("scoring")

    ws = wb.active
    ws.title = "Synthese"
    _header_row(ws, ["Champ", "Valeur"])
    rows = [
        ("document_kind", extraction.get("document_kind")),
        ("source_filename", extraction.get("source_filename")),
        ("pages_processed", extraction.get("pages_processed")),
        ("completeness_pct", extraction.get("completeness_pct")),
        ("processing_time_ms", extraction.get("processing_time_ms")),
        ("ocr_method", extraction.get("ocr_method") or extraction.get("method")),
        ("summary", extraction.get("summary")),
    ]
    if scoring and scoring.get("decision"):
        decision = scoring["decision"]
        rows.extend(
            [
                ("score", decision.get("score")),
                ("classe", decision.get("classe")),
                ("decision", decision.get("decision")),
                ("recommandation", decision.get("recommandation")),
                ("blocking_status", decision.get("blocking_status")),
            ]
        )
    if payload.get("scoring_warning"):
        rows.append(("scoring_warning", payload["scoring_warning"]))
    if payload.get("scoring_skipped_reason"):
        rows.append(("scoring_skipped_reason", payload["scoring_skipped_reason"]))

    for i, (key, value) in enumerate(rows, start=2):
        ws.cell(i, 1, key)
        ws.cell(i, 2, _cell(value))
    _autofit(ws)

    ws_el = wb.create_sheet("Elements_PCGM")
    _header_row(ws_el, ["Code", "Label", "Section", "Valeur", "Statut"])
    elements = extraction.get("elements") or []
    for i, el in enumerate(elements, start=2):
        ws_el.cell(i, 1, el.get("code"))
        ws_el.cell(i, 2, el.get("label"))
        ws_el.cell(i, 3, el.get("section") or el.get("source"))
        ws_el.cell(i, 4, el.get("value"))
        ws_el.cell(i, 5, el.get("detection_status") or el.get("status"))
    _autofit(ws_el)

    ws_in = wb.create_sheet("Donnees_financieres")
    _header_row(ws_in, ["Champ", "Valeur"])
    scoring_input = extraction.get("scoring_input") or {}
    for i, (key, value) in enumerate(sorted(scoring_input.items()), start=2):
        ws_in.cell(i, 1, key)
        ws_in.cell(i, 2, value)
    _autofit(ws_in)

    if scoring:
        ws_r = wb.create_sheet("Ratios")
        _header_row(
            ws_r,
            ["Clé", "Label", "Valeur", "Unité", "Statut", "Seuil", "Formule", "Raison"],
        )
        ratios = scoring.get("ratios") or {}
        if isinstance(ratios, dict):
            for i, (key, detail) in enumerate(ratios.items(), start=2):
                ws_r.cell(i, 1, key)
                ws_r.cell(i, 2, detail.get("label"))
                ws_r.cell(i, 3, detail.get("value"))
                ws_r.cell(i, 4, detail.get("unit"))
                ws_r.cell(i, 5, detail.get("status"))
                ws_r.cell(i, 6, detail.get("threshold"))
                ws_r.cell(i, 7, detail.get("formula"))
                ws_r.cell(i, 8, detail.get("reason"))
        _autofit(ws_r)

        ws_a = wb.create_sheet("Axes")
        _header_row(ws_a, ["Axe", "Score", "Pondération", "Contribution"])
        for i, (name, key) in enumerate(
            [
                ("Axe 1 — Financier", "axe1"),
                ("Axe 2 — Comportemental", "axe2"),
                ("Axe 3 — Sectoriel", "axe3"),
            ],
            start=2,
        ):
            axe = scoring.get(key) or {}
            ws_a.cell(i, 1, name)
            ws_a.cell(i, 2, axe.get("score"))
            ws_a.cell(i, 3, axe.get("ponderation"))
            ws_a.cell(i, 4, axe.get("contribution"))
        _autofit(ws_a)


def build_excel_bytes(payload: dict[str, Any]) -> tuple[bytes, str]:
    """Construit un classeur Excel (liasse legacy ou analyse financière PDF)."""
    wb = Workbook()

    # Lot multi-documents
    documents = payload.get("documents")
    if isinstance(documents, list) and documents:
        # Feuille index
        ws_idx = wb.active
        ws_idx.title = "Index"
        _header_row(ws_idx, ["#", "Fichier", "Score", "Classe", "Décision", "Statut"])
        for i, doc in enumerate(documents, start=2):
            extraction = doc.get("extraction") or {}
            analysis = doc.get("analysis") or {}
            decision = analysis.get("decision") or {}
            ws_idx.cell(i, 1, i - 1)
            ws_idx.cell(i, 2, extraction.get("source_filename") or doc.get("filename"))
            ws_idx.cell(i, 3, _cell(decision.get("score") or analysis.get("final_score")))
            ws_idx.cell(i, 4, decision.get("risk_class"))
            ws_idx.cell(i, 5, decision.get("decision"))
            ws_idx.cell(i, 6, doc.get("status") or ("ok" if analysis else "extract"))
        _autofit(ws_idx)

        for idx, doc in enumerate(documents, start=1):
            extraction = doc.get("extraction") or {}
            analysis = doc.get("analysis")
            if analysis:
                _write_financial_analysis_sheets(
                    wb,
                    extraction=extraction,
                    analysis=analysis,
                    sheet_prefix=f"D{idx}_",
                )
            elif extraction.get("elements") or extraction.get("scoring_input"):
                # legacy single embedded — rare in batch
                continue
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue(), f"analyse_financiere_lot_{_ts()}.xlsx"

    # Document unique analyse financière
    if payload.get("analysis") or (
        isinstance(payload.get("extraction"), dict)
        and payload.get("extraction", {}).get("pages") is not None
        and "dataset" in (payload.get("analysis") or {})
    ):
        extraction = payload.get("extraction") or {}
        analysis = payload.get("analysis") or {}
        # Remplace feuille vide
        default = wb.active
        wb.remove(default)
        _write_financial_analysis_sheets(
            wb,
            extraction=extraction,
            analysis=analysis,
        )
        # Assure au moins une feuille
        if not wb.worksheets:
            wb.create_sheet("Synthese")
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue(), f"analyse_financiere_{_ts()}.xlsx"

    # Format liasse historique
    _write_legacy_liasse_sheets(wb, payload)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), f"liasse_resultats_{_ts()}.xlsx"
